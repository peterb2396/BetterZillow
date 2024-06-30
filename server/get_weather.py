import requests
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
import os
import signal
import sys
from datetime import datetime, timedelta

# Function to call the weather API and get the data
def get_weather_data(lat, lon):
    url = f"https://archive-api.open-meteo.com/v1/archive?latitude={lat}&longitude={lon}&start_date=2022-01-01&end_date=2022-12-31&daily=temperature_2m_max,temperature_2m_min,precipitation_sum,wind_speed_10m_max&temperature_unit=fahrenheit&wind_speed_unit=mph&precipitation_unit=inch"
    response = requests.get(url)
    data = response.json()
    return data

# Function to remove outliers using the IQR method
def remove_outliers(data):
    if not isinstance(data, pd.Series):
        data = pd.Series(data)
    
    q1 = data.quantile(0.25)
    q3 = data.quantile(0.75)
    iqr = q3 - q1
    
    filtered_data = data[~((data < (q1 - 1.5 * iqr)) | (data > (q3 + 1.5 * iqr)))]
    
    sorted_data = filtered_data.sort_values()
    
    trim = int(0.03 * len(data))
    final_data = sorted_data.iloc[trim:-trim]
    
    return final_data

# Function to extract the required weather information from the API response
def extract_weather_info(data):
    daily = data['daily']
    
    precipitation_sum = pd.Series(daily['precipitation_sum'])
    temperature_2m_max = pd.Series(daily['temperature_2m_max'])
    temperature_2m_min = pd.Series(daily['temperature_2m_min'])
    wind_speed_10m_max = pd.Series(daily['wind_speed_10m_max'])
    
    precipitation_sum = remove_outliers(precipitation_sum)
    temperature_2m_max = remove_outliers(temperature_2m_max)
    temperature_2m_min = remove_outliers(temperature_2m_min)
    wind_speed_10m_max = remove_outliers(wind_speed_10m_max)
    
    total_rainfall = precipitation_sum.sum()
    average_daily_rainfall = total_rainfall / len(precipitation_sum)
    average_wind_speed = wind_speed_10m_max.mean()
    max_temperature = temperature_2m_max.max()
    min_temperature = temperature_2m_max.min()
    
    return total_rainfall, average_daily_rainfall, average_wind_speed, max_temperature, min_temperature

def get_sleep_duration(error_reason):
    now = datetime.now()
    
    if 'Minutely' in error_reason:
        next_minute = now.replace(second=0, microsecond=0) + timedelta(minutes=1)
        sleep_duration = (next_minute - now).total_seconds() +60
        print(f"Retrying in {int(sleep_duration)} seconds...")
        return sleep_duration
    
    elif 'Hourly' in error_reason:
        next_hour = now.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
        sleep_duration = (next_hour - now).total_seconds() +60
        print(f"Retrying in {int(sleep_duration)} seconds...")
        return sleep_duration
    
    elif 'Daily' in error_reason:
        next_day = now.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
        sleep_duration = (next_day - now).total_seconds() +60
        print(f"Retrying in {int(sleep_duration)} seconds...")
        return sleep_duration
    
    else:
        return 5
# Function to fetch and process weather data for a city
def process_city(row):
    lat = row['lat']
    lon = row['lng']
    population = row['population']
    density = row['density']
    city = row['city']
    state_name = row['state_name']
    id = row['state_id']
    weather_data = None
    
    while True:
        try:
            weather_data = get_weather_data(lat, lon)
            print(f"Processed {city}, {state_name}")
            return id, city, state_name, lat, lon, population, density, extract_weather_info(weather_data)
        except Exception as e:
            error_reason = 'Unknown error' if not weather_data else weather_data.get('reason', 'Unknown error')
            print(f"Error processing {city}, {state_name}: {error_reason}")
            
            sleep_duration = get_sleep_duration(error_reason)
            
            if sleep_duration:
                time.sleep(sleep_duration)
            else:
                break

# Function to handle SIGINT and save the workbook
def signal_handler(sig, frame):
    print("SIGINT received. Saving the workbook and exiting...")
    wb.save("weather.xlsx")
    sys.exit(0)

# Register the signal handler for SIGINT
signal.signal(signal.SIGINT, signal_handler)

# Read the original Excel file
df = pd.read_excel('cities.xlsx')

# Where to start adding rows. If new excel sheet, start at 2 (1 after header)
start = 2

# Check if 'weather.xlsx' exists and load or create the workbook
if os.path.exists("weather.xlsx"):
    wb = load_workbook(filename="weather.xlsx")
    ws = wb.active
    start = ws.max_row + 1
else:
    wb = Workbook()
    ws = wb.active
    # Add headers for the new columns if the file is newly created
    new_columns = ['id', 'State', 'City ', 'Latitude', 'Longitude', 'Total Rainfall', 'Average Daily Rainfall', 'Average Wind Speed', 'Maximum Temperature', 'Minimum Temperature']
    for col_num, header in enumerate(new_columns, start=1):
        ws.cell(row=1, column=col_num, value=header)

# Process each city sequentially
try:
    for index, row in df.iloc[start - 2:].iterrows():
        id, city, state_name, lat, lon, population, density, weather_info = process_city(row)
        total_rainfall, average_daily_rainfall, average_wind_speed, max_temperature, min_temperature = weather_info

        # Update the Excel sheet with the new data
        ws.cell(row=index+2, column=1, value=id)
        ws.cell(row=index+2, column=2, value=state_name)
        ws.cell(row=index+2, column=3, value=city)
        ws.cell(row=index+2, column=4, value=lat)
        ws.cell(row=index+2, column=5, value=lon)
        ws.cell(row=index+2, column=6, value=population)
        ws.cell(row=index+2, column=7, value=density)
        ws.cell(row=index+2, column=8, value=total_rainfall)
        ws.cell(row=index+2, column=9, value=average_daily_rainfall)
        ws.cell(row=index+2, column=10, value=average_wind_speed)
        ws.cell(row=index+2, column=11, value=max_temperature)
        ws.cell(row=index+2, column=12, value=min_temperature)

        # Save the workbook after updating each row
        wb.save('weather.xlsx')
finally:
    # Ensure the workbook is saved if the loop is interrupted
    wb.save("weather.xlsx")
    print("Workbook saved successfully.")
